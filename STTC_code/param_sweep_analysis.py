# param_sweep_analysis.py
# BRC R1[2] STTC 参数扫描分析（Python 复现，指标口径与 BCT/MATLAB 逐位对齐，已经工作流三套独立实现 + 机器精度核验）
#
# 产出（写到 <merged>\param_sweep\ 下，全部 .xlsx，分块版式对齐你的 metrics_3point_final.xlsx）：
#   1) metrics_baseline.xlsx          —— thr=0.35 / kS=mean_nz / node_3rd=cloc，用于和你的截图逐数核对（保真性自检）
#   2) metrics_normalize_meannz.xlsx  —— 任务1：每盘 W1 先除以自身非零边均值（抹掉整体连接强度）再算 CC/CPL/Eglob/Hub
#   3) metrics_threshold_sweep.xlsx   —— 任务4：thr 取 0.10/0.20/0.35/0.50 各一个 sheet + 残差随阈值汇总 sheet
#   4) metrics_hub_bc.xlsx            —— 任务5：Hub 第三指标用 betweenness(bc，对缩放严格不变)重算
#   5) edge_category_proportions.xlsx —— 任务3：四类边(增强/减弱/新增/修剪)计数 + 占比，逐盘 + 合计
#
# 运行：python param_sweep_analysis.py      （需 numpy/scipy/pandas/openpyxl，无需 MATLAB）
import os, heapq
import numpy as np
import scipy.io as sio
from scipy.sparse.csgraph import shortest_path
import pandas as pd
from openpyxl import Workbook

MERGED = r'E:\Recognition-in-Biological-Neural-Networks\Data\merged'
OUTDIR = os.path.join(MERGED, 'param_sweep')
os.makedirs(OUTDIR, exist_ok=True)

# 9 盘 pre/post1/post2，照抄 main_edge_change.m cfg.designation
DESIGNATION = {
    '0117-1': (1, [2, 6]), '0206-1': (2, [4, 5]), '0206-2': (3, [4, 5]),
    '0206-3': (2, [3, 5]), '0206-4': (1, [3, 5]), '0928-3': (2, [3, 6]),
    '1108-1': (2, [3, 6]), '1213-5': (3, [4, 5]), '1227-1': (3, [4, 6]),
}
ORDER = ['0117-1','0206-1','0206-2','0206-3','0206-4','0928-3','1108-1','1213-5','1227-1']
HUB_TOP = 0.40

# ---------------- 数据加载 ----------------
def load_full(cell, s):
    f = os.path.join(MERGED, cell, 'sttc', f'sttc-spikes_spon{s}.mat')
    d = sio.loadmat(f)
    A = np.array(d['adjM'], float)
    A[np.isnan(A)] = 0.0
    A[A < 0] = 0.0
    active = None
    if 'activeElectrode' in d and np.size(d['activeElectrode']) > 0:
        active = np.array(d['activeElectrode']).ravel().astype(int)
    return A, active

def w1_subgraph(A_full, thr):
    # 口径 = main_network_scaling_control.m w1_subgraph：清负(已做)→剔零→<thr=0→再剔零
    W0 = A_full.copy()
    zr = ~(W0 == 0).all(1); zc = ~(W0 == 0).all(0)
    Wa = W0[np.ix_(zr, zc)]
    Wt = Wa.copy(); Wt[Wt < thr] = 0.0
    zr2 = ~(Wt == 0).all(1); zc2 = ~(Wt == 0).all(0)
    return Wt[np.ix_(zr2, zc2)]

def mean_nz_full(A_full, thr):
    W = A_full.copy(); W[W < thr] = 0.0
    N = W.shape[0]; off = ~np.eye(N, dtype=bool)
    v = W[off]; v = v[v > 0]
    return v.mean() if v.size else 0.0

# ---------------- 距离 / BCT 指标 ----------------
def dist_mat(W):
    L = np.full_like(W, np.inf); m = W > 0; L[m] = 1.0 / W[m]; np.fill_diagonal(L, 0.0)
    return shortest_path(L, method='D', directed=False)

def cc_onnela(W):
    K = (W != 0).sum(1).astype(float)
    A3 = np.cbrt(W); cyc3 = np.diag(A3 @ A3 @ A3)
    with np.errstate(divide='ignore', invalid='ignore'):
        C = np.where(cyc3 > 0, cyc3 / (K * (K - 1)), 0.0)
    pos = C > 0
    return C[pos].mean() if pos.any() else 0.0

def eglob(W):
    D = dist_mat(W); off = ~np.eye(W.shape[0], dtype=bool)
    inv = np.zeros_like(D); pos = (D > 0) & np.isfinite(D); inv[pos] = 1.0 / D[pos]
    return inv[off].mean()

def cpl(W):
    N = W.shape[0]; D = dist_mat(W); off = ~np.eye(N, dtype=bool)
    if np.isinf(D[off]).any():
        return np.nan
    return D.sum() / (N * (N - 1))

def cloc_node(W):
    D = dist_mat(W); s = D.sum(1)
    c = np.where(s > 0, 1.0 / s, 0.0); c[~np.isfinite(c)] = 0.0
    return c

def betweenness_node(W):
    # Brandes(加权,Dijkstra) on L=1/W；口径同 betweenness_wei(L)。对全局缩放严格不变。
    L = np.full_like(W, np.inf); m = W > 0; L[m] = 1.0 / W[m]
    n = W.shape[0]
    nbrs = [np.where((W[i] > 0))[0] for i in range(n)]
    BC = np.zeros(n)
    for s in range(n):
        S = []; P = [[] for _ in range(n)]; sigma = np.zeros(n); sigma[s] = 1.0
        dist = np.full(n, np.inf); dist[s] = 0.0; done = np.zeros(n, bool)
        Q = [(0.0, s)]
        while Q:
            d, v = heapq.heappop(Q)
            if done[v]:
                continue
            done[v] = True; S.append(v)
            for w in nbrs[v]:
                vw = d + L[v, w]
                if vw < dist[w] - 1e-15:
                    dist[w] = vw; sigma[w] = sigma[v]; P[w] = [v]; heapq.heappush(Q, (vw, w))
                elif abs(vw - dist[w]) <= 1e-12:
                    sigma[w] += sigma[v]; P[w].append(v)
        delta = np.zeros(n)
        while S:
            w = S.pop()
            for v in P[w]:
                if sigma[w] > 0:
                    delta[v] += (sigma[v] / sigma[w]) * (1.0 + delta[w])
            if w != s:
                BC[w] += delta[w]
    return BC

def eloc_case2(W):
    # efficiency_wei(W,2) 局部效率，照抄 BCT case 2
    n = W.shape[0]; A = (W > 0).astype(float)
    L = W.copy(); mk = W > 0; L[mk] = 1.0 / L[mk]
    cbrt_W = np.cbrt(W); cbrt_L = np.cbrt(L)
    E = np.zeros(n)
    for u in range(n):
        V = np.where((A[u, :] > 0) | (A[:, u] > 0))[0]
        if V.size < 2:
            continue
        sw = cbrt_W[u, V] + cbrt_W[V, u]
        sub = cbrt_L[np.ix_(V, V)].copy()
        M = np.where(sub > 0, sub, np.inf); np.fill_diagonal(M, 0.0)
        D = shortest_path(M, method='D', directed=False)
        di = np.zeros_like(D); pos = (D > 0) & np.isfinite(D); di[pos] = 1.0 / D[pos]; np.fill_diagonal(di, 0.0)
        se = di + di.T
        numer = np.sum(np.outer(sw, sw) * se) / 2.0
        if numer != 0:
            sa = A[u, V] + A[V, u]
            denom = sa.sum() ** 2 - (sa ** 2).sum()
            if denom != 0:
                E[u] = numer / denom
    return E

# ---------------- 节点指标 + hub ----------------
def node_metrics(W):
    str_ = W.sum(1); deg = (W > 0).sum(1).astype(float)
    mean_str = np.where(deg > 0, str_ / np.maximum(deg, 1e-12), 0.0)
    return dict(mean_str=mean_str, eloc=eloc_case2(W), cloc=cloc_node(W), bc=betweenness_node(W))

def hub_score(target, pool, third):
    aS = np.concatenate([p['mean_str'] for p in pool])
    aE = np.concatenate([p['eloc'] for p in pool])
    a3 = np.concatenate([p[third] for p in pool])
    def thr_of(a):
        g = int(min(max(round(HUB_TOP * len(a)), 1), len(a)))
        return np.sort(a)[::-1][g - 1]
    thS, thE, th3 = thr_of(aS), thr_of(aE), thr_of(a3)
    h = (target['mean_str'] >= thS).astype(int) + (target['eloc'] >= thE).astype(int) + (target[third] >= th3).astype(int)
    return h.sum() / len(target['mean_str'])

# ---------------- 单盘指标计算 ----------------
def compute_dish(cell, thr, normalize=False, third='cloc'):
    pre, posts = DESIGNATION[cell]
    stages = [pre] + posts
    W1 = {}; mnz_full = {}; nm = {}
    for s in stages:
        A, _ = load_full(cell, s)
        w = w1_subgraph(A, thr)
        mnz_full[s] = mean_nz_full(A, thr)
        if normalize:
            mloc = w[w > 0].mean() if (w > 0).any() else 1.0   # 除以自身非零边均值
            w = w / mloc if mloc > 0 else w
        W1[s] = w
        nm[s] = node_metrics(w)
    pool = [nm[s] for s in stages]   # hub 阈值池 = 三阶段拼接（口径同 hub_score basePool）

    out = dict(cell=cell, pre=pre, post1=posts[0], post2=posts[1])
    Wpre = W1[pre]
    for metric, fn in [('CC', cc_onnela), ('CPL', cpl), ('Eglob', eglob)]:
        out[metric + '_pre'] = fn(Wpre)
    out['Hub_pre'] = hub_score(nm[pre], pool, third)
    out['Hub_bc_pre'] = hub_score(nm[pre], pool, 'bc')   # 并行 bc 口径 hub(第三指标=介数)
    out['meanSTTC_pre'] = mnz_full[pre]

    for k, t in enumerate(posts, start=1):
        kS = mnz_full[t] / mnz_full[pre] if mnz_full[pre] > 0 else np.nan
        if normalize:
            scaled = Wpre.copy()          # 归一化空间里缩放对照 = pre 本身（整体标量被抹掉）
        else:
            scaled = Wpre * kS
        out[f'kS_p{k}'] = kS
        out[f'meanSTTC_p{k}'] = mnz_full[t]
        for metric, fn in [('CC', cc_onnela), ('CPL', cpl), ('Eglob', eglob)]:
            act = fn(W1[t]); scl = fn(scaled)
            out[f'{metric}_p{k}_act'] = act
            out[f'{metric}_p{k}_scl'] = scl
            out[f'{metric}_p{k}_res'] = act - scl
        nm_scaled = node_metrics(scaled)
        hub_act = hub_score(nm[t], pool, third)
        hub_scl = hub_score(nm_scaled, pool, third)
        out['Hub_p%d_act' % k] = hub_act
        out['Hub_p%d_scl' % k] = hub_scl
        out['Hub_p%d_res' % k] = hub_act - hub_scl
        hb_act = hub_score(nm[t], pool, 'bc'); hb_scl = hub_score(nm_scaled, pool, 'bc')
        out['Hub_bc_p%d_act' % k] = hb_act
        out['Hub_bc_p%d_scl' % k] = hb_scl
        out['Hub_bc_p%d_res' % k] = hb_act - hb_scl
    return out

# ---------------- Excel 分块版式（对齐 metrics_3point_final.xlsx）----------------
BLOCK_START = {'CC': 1, 'CPL': 12, 'Eglob': 24, 'Hub': 36}  # 各指标块起始 Excel 行
def write_metrics_sheet(ws, rows, normalized=False):
    scl_label = '缩放(=pre)' if normalized else '缩放'
    # CC 块带 盘/阶段/meanSTTC/kS 列
    h = ['盘', 'pre号', 'post1号', 'post2号', '', 'CC_pre',
         'CC_p1实测', 'CC_p1' + scl_label, 'CC_p1残差', 'CC_p2实测', 'CC_p2' + scl_label, 'CC_p2残差', '',
         'meanSTTC_pre', 'meanSTTC_p1', 'meanSTTC_p2', '', 'kS_p1', 'kS_p2']
    for c, v in enumerate(h, 1):
        ws.cell(1, c, v)
    for i, r in enumerate(rows):
        rr = 2 + i
        ws.cell(rr, 1, r['cell']); ws.cell(rr, 2, r['pre']); ws.cell(rr, 3, r['post1']); ws.cell(rr, 4, r['post2'])
        ws.cell(rr, 6, rnd(r['CC_pre']))
        ws.cell(rr, 7, rnd(r['CC_p1_act'])); ws.cell(rr, 8, rnd(r['CC_p1_scl'])); ws.cell(rr, 9, rnd(r['CC_p1_res']))
        ws.cell(rr, 10, rnd(r['CC_p2_act'])); ws.cell(rr, 11, rnd(r['CC_p2_scl'])); ws.cell(rr, 12, rnd(r['CC_p2_res']))
        ws.cell(rr, 14, rnd(r['meanSTTC_pre'])); ws.cell(rr, 15, rnd(r['meanSTTC_p1'])); ws.cell(rr, 16, rnd(r['meanSTTC_p2']))
        ws.cell(rr, 18, rnd(r['kS_p1'])); ws.cell(rr, 19, rnd(r['kS_p2']))
    # CPL/Eglob/Hub 块（仅 F-L 七列）
    for metric in ['CPL', 'Eglob', 'Hub']:
        r0 = BLOCK_START[metric]
        hh = [f'{metric}_pre', f'{metric}_p1实测', f'{metric}_p1' + scl_label, f'{metric}_p1残差',
              f'{metric}_p2实测', f'{metric}_p2' + scl_label, f'{metric}_p2残差']
        for c, v in enumerate(hh):
            ws.cell(r0, 6 + c, v)
        for i, r in enumerate(rows):
            rr = r0 + 1 + i
            ws.cell(rr, 6, rnd(r[f'{metric}_pre']))
            ws.cell(rr, 7, rnd(r[f'{metric}_p1_act'])); ws.cell(rr, 8, rnd(r[f'{metric}_p1_scl'])); ws.cell(rr, 9, rnd(r[f'{metric}_p1_res']))
            ws.cell(rr, 10, rnd(r[f'{metric}_p2_act'])); ws.cell(rr, 11, rnd(r[f'{metric}_p2_scl'])); ws.cell(rr, 12, rnd(r[f'{metric}_p2_res']))

def rnd(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    return round(float(x), 4)

def build_rows(thr, normalize=False, third='cloc'):
    return [compute_dish(c, thr, normalize, third) for c in ORDER]

# ---------------- 任务3：四类边占比 ----------------
def classify_edges(Wpre, Wpost, common, tau=0.35, delta=0.10):
    idx = np.array(common, int) - 1
    a = Wpre[np.ix_(idx, idx)]; b = Wpost[np.ix_(idx, idx)]; m = idx.size
    cnt = dict(enhanced=0, weakened=0, new=0, pruned=0, stable=0)
    for ii in range(m - 1):
        for jj in range(ii + 1, m):
            av, bv = a[ii, jj], b[ii, jj]
            ppre, ppost = av >= tau, bv >= tau
            if not ppre and not ppost:
                continue
            if ppre and ppost:
                if bv - av > delta: cnt['enhanced'] += 1
                elif av - bv > delta: cnt['weakened'] += 1
                else: cnt['stable'] += 1
            elif (not ppre) and ppost and bv >= tau + delta: cnt['new'] += 1
            elif ppre and (not ppost) and av >= tau + delta: cnt['pruned'] += 1
            else: cnt['stable'] += 1
    return cnt

def load_full_symm(cell, s):
    A, active = load_full(cell, s)
    A = (A + A.T) / 2.0; np.fill_diagonal(A, 0.0)
    if active is None:
        active = np.where((A > 0).any(1))[0] + 1
    return A, active

def edge_proportions():
    rows = []
    for cell in ORDER:
        pre, posts = DESIGNATION[cell]
        Wpre, actPre = load_full_symm(cell, pre)
        for k, t in enumerate(posts, 1):
            Wpost, actPost = load_full_symm(cell, t)
            common = np.intersect1d(actPre, actPost)
            common = common[(common >= 1) & (common <= Wpre.shape[0])]
            c = classify_edges(Wpre, Wpost, common)
            changed = c['enhanced'] + c['weakened'] + c['new'] + c['pruned']
            mag = c['enhanced'] + c['weakened']; struc = c['new'] + c['pruned']
            rows.append(dict(盘=cell, 对比=f'pre→post{k}(spon{pre}→{t})', 公共电极=int(common.size),
                             增强=c['enhanced'], 减弱=c['weakened'], 新增=c['new'], 修剪=c['pruned'], 变化边合计=changed,
                             增强占比=pct(c['enhanced'], changed), 减弱占比=pct(c['weakened'], changed),
                             新增占比=pct(c['new'], changed), 修剪占比=pct(c['pruned'], changed),
                             幅值类占比=pct(mag, changed), 结构类占比=pct(struc, changed)))
    df = pd.DataFrame(rows)
    tot = dict(盘='合计(9盘18对)', 对比='', 公共电极=int(df['公共电极'].sum()),
               增强=int(df['增强'].sum()), 减弱=int(df['减弱'].sum()), 新增=int(df['新增'].sum()),
               修剪=int(df['修剪'].sum()), 变化边合计=int(df['变化边合计'].sum()))
    ch = tot['变化边合计']
    tot.update(增强占比=pct(tot['增强'], ch), 减弱占比=pct(tot['减弱'], ch), 新增占比=pct(tot['新增'], ch),
               修剪占比=pct(tot['修剪'], ch), 幅值类占比=pct(tot['增强'] + tot['减弱'], ch),
               结构类占比=pct(tot['新增'] + tot['修剪'], ch))
    return pd.concat([df, pd.DataFrame([tot])], ignore_index=True)

def pct(x, tot):
    return round(100.0 * x / tot, 1) if tot else 0.0

# ==================== 主流程 ====================
def main():
    # 1) baseline 自检 + 写出
    base_rows = build_rows(0.35, normalize=False, third='cloc')
    wb = Workbook(); write_metrics_sheet(wb.active, base_rows); wb.active.title = 'metrics'
    wb.save(os.path.join(OUTDIR, 'metrics_baseline.xlsx'))
    print('=== baseline 自检（与截图逐数核对，0117-1 / 1108-1）===')
    for r in base_rows:
        if r['cell'] in ('0117-1', '1108-1'):
            print(f"{r['cell']} CC_pre={r['CC_pre']:.4f} CC_p1act={r['CC_p1_act']:.4f} CC_p1scl={r['CC_p1_scl']:.4f} "
                  f"CC_p1res={r['CC_p1_res']:.4f} kS1={r['kS_p1']:.4f} meanpre={r['meanSTTC_pre']:.4f} "
                  f"Hub_pre={r['Hub_pre']:.4f} Hub_p1act={r['Hub_p1_act']:.4f} Hub_p1scl={r['Hub_p1_scl']:.4f}")

    # 2) 任务1：归一化（除以非零边均值）
    norm_rows = build_rows(0.35, normalize=True, third='cloc')
    wb = Workbook(); write_metrics_sheet(wb.active, norm_rows, normalized=True); wb.active.title = 'norm_meannz'
    wb.save(os.path.join(OUTDIR, 'metrics_normalize_meannz.xlsx'))
    print('\n=== 任务1 归一化后 残差(post−pre, 已抹掉整体强度) 概览 ===')
    summarize_res(norm_rows)

    # 4) 任务4：阈值扫描
    thr_list = [0.10, 0.20, 0.35, 0.50]
    wb = Workbook(); first = True; sweep_summary = []
    for thr in thr_list:
        rows = build_rows(thr, normalize=False, third='cloc')
        ws = wb.active if first else wb.create_sheet()
        ws.title = f'thr_{thr:.2f}'; first = False
        write_metrics_sheet(ws, rows)
        # 残差汇总
        s = res_stats(rows)
        s['thr'] = thr
        s['W1平均密度'] = round(np.mean([dish_density(c, thr) for c in ORDER]), 3)
        sweep_summary.append(s)
        print(f'\n=== 任务4 thr={thr:.2f}  W1平均密度={s["W1平均密度"]:.3f}  '
              f'CC残差均|·|={s["CC_mean_abs_res"]:.4f} Eglob残差均|·|={s["Eglob_mean_abs_res"]:.4f} '
              f'CPL残差均|·|={s["CPL_mean_abs_res"]:.4f} Hub残差均={s["Hub_mean_res"]:.3f}')
    ws = wb.create_sheet('残差随阈值汇总')
    dfsum = pd.DataFrame(sweep_summary)[['thr', 'W1平均密度', 'CC_mean_abs_res', 'CC_mean_res',
                                          'Eglob_mean_abs_res', 'CPL_mean_abs_res', 'Hub_mean_res', 'kS均值']]
    for c, name in enumerate(dfsum.columns, 1):
        ws.cell(1, c, name)
    for i, (_, row) in enumerate(dfsum.iterrows(), 2):
        for c, name in enumerate(dfsum.columns, 1):
            ws.cell(i, c, round(float(row[name]), 4))
    wb.save(os.path.join(OUTDIR, 'metrics_threshold_sweep.xlsx'))

    # 5) 任务5：Hub 用 bc
    bc_rows = build_rows(0.35, normalize=False, third='bc')
    wb = Workbook(); write_metrics_sheet(wb.active, bc_rows); wb.active.title = 'metrics_hub_bc'
    wb.save(os.path.join(OUTDIR, 'metrics_hub_bc.xlsx'))
    print('\n=== 任务5 Hub(第三指标=bc) 残差(实测−缩放) 逐盘 ===')
    for r in bc_rows:
        print(f"{r['cell']}  Hub_p1: 实测{r['Hub_p1_act']:.3f} 缩放{r['Hub_p1_scl']:.3f} 残差{r['Hub_p1_res']:+.3f}  | "
              f"Hub_p2: 实测{r['Hub_p2_act']:.3f} 缩放{r['Hub_p2_scl']:.3f} 残差{r['Hub_p2_res']:+.3f}")

    # 3) 任务3：四类边占比
    df = edge_proportions()
    df.to_excel(os.path.join(OUTDIR, 'edge_category_proportions.xlsx'), index=False, sheet_name='四类占比')
    print('\n=== 任务3 四类边占比（合计行）===')
    print(df.iloc[-1].to_string())

    print('\n全部 Excel 已写到：', OUTDIR)

def dish_density(cell, thr):
    A, _ = load_full(cell, DESIGNATION[cell][0]); w = w1_subgraph(A, thr); N = w.shape[0]
    return (w > 0).sum() / (N * (N - 1)) if N > 1 else 0.0

def res_stats(rows):
    def col(metric, suf):
        v = []
        for r in rows:
            for k in (1, 2):
                x = r[f'{metric}_p{k}_{suf}']
                if x is not None and not (isinstance(x, float) and np.isnan(x)):
                    v.append(x)
        return np.array(v)
    return dict(
        CC_mean_abs_res=np.mean(np.abs(col('CC', 'res'))),
        CC_mean_res=np.mean(col('CC', 'res')),
        Eglob_mean_abs_res=np.mean(np.abs(col('Eglob', 'res'))),
        CPL_mean_abs_res=np.nanmean(np.abs(col('CPL', 'res'))),
        Hub_mean_res=np.mean(col('Hub', 'res')),
        kS均值=np.mean([r[f'kS_p{k}'] for r in rows for k in (1, 2)]),
    )

def summarize_res(rows):
    s = res_stats(rows)
    print(f"  CC  残差(post−pre) 均值={s['CC_mean_res']:+.4f}  均|·|={s['CC_mean_abs_res']:.4f}")
    print(f"  Eglob 残差 均|·|={s['Eglob_mean_abs_res']:.4f}   CPL 残差 均|·|={s['CPL_mean_abs_res']:.4f}")
    # 逐盘 CC 残差符号
    pos = sum(1 for r in rows for k in (1, 2) if r[f'CC_p{k}_res'] > 0)
    neg = sum(1 for r in rows for k in (1, 2) if r[f'CC_p{k}_res'] < 0)
    print(f"  CC 残差符号：正{pos} / 负{neg}（共18对）")

if __name__ == '__main__':
    main()
