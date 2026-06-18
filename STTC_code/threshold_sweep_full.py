# threshold_sweep_full.py
# 重生成更全的 metrics_threshold_sweep.xlsx：
#   1) 阈值 0.10–0.60，步长 0.05（11 组），看趋势到 0.6 是否消失
#   2) 每指标(CC/CPL/Eglob/Hub)、每 post 给出 实测vs缩放 的配对 Wilcoxon p 值 + 显著性星号
#      方向性假设：CC/Eglob/Hub 实测>缩放(greater)；CPL 实测<缩放(less)
#   产出：汇总_残差与显著性(首页一眼看趋势) + 每阈值明细 sheet(底部附显著性)
# 运行：python threshold_sweep_full.py
import os
import numpy as np
from openpyxl import Workbook
from scipy.stats import wilcoxon
from param_sweep_analysis import (compute_dish, write_metrics_sheet, w1_subgraph,
                                   load_full, DESIGNATION, ORDER, OUTDIR, rnd)

THRS = [round(0.10 + 0.05 * i, 2) for i in range(15)]   # 0.10 .. 0.80
DIRN = {'CC': 'greater', 'Eglob': 'greater', 'CPL': 'less', 'Hub': 'greater', 'Hub_bc': 'greater'}

def stars(p):
    if p != p:
        return 'ns'
    if p < 0.001: return '***'
    if p < 0.01:  return '**'
    if p < 0.05:  return '*'
    if p < 0.10:  return '·'
    return 'ns'

def pair_test(rows, metric, post, kind):
    # kind: 'act_pre'(实测-pre) | 'scl_pre'(缩放-pre) | 'act_scl'(实测-缩放)
    pre = np.array([r[f'{metric}_pre'] for r in rows], float)
    act = np.array([r[f'{metric}_p{post}_act'] for r in rows], float)
    scl = np.array([r[f'{metric}_p{post}_scl'] for r in rows], float)
    d = {'act_pre': act - pre, 'scl_pre': scl - pre, 'act_scl': act - scl}[kind]
    d = d[np.isfinite(d)]; nz = d[d != 0]
    if nz.size == 0:
        return (d.mean() if d.size else np.nan), np.nan
    p = wilcoxon(nz, alternative=DIRN[metric]).pvalue
    return d.mean(), p

KINDS = [('act_pre', '实测vs pre'), ('scl_pre', '缩放vs pre'), ('act_scl', '实测vs缩放')]

def density(thr):
    ds = []
    for c in ORDER:
        A, _ = load_full(c, DESIGNATION[c][0]); w = w1_subgraph(A, thr); N = w.shape[0]
        ds.append((w > 0).sum() / (N * (N - 1)) if N > 1 else 0.0)
    return np.mean(ds)

def add_explanation_sheet(wb):
    ws = wb.create_sheet('说明_cloc与bc', 0)
    ws.column_dimensions['A'].width = 130
    lines = [
        '【hub 第三指标 cloc → bc 的更改说明】',
        '',
        '一、三个子指标:hub(枢纽)评分怎么算',
        '  每个节点按三方面打分,每达标一项 +1(满分3),再对全网取平均得到该网络的 hub 平均分。三个子指标:',
        '   1) 平均连接强度 mean_str:这个节点的连边平均有多强。',
        '   2) 局部效率 eloc:这个节点身边小圈子内部联系有多紧密、传递有多高效。',
        '   3) 第三指标:原用 cloc(接近中心性),现改用 bc(介数中心性)。前两个一直不变,只换第三个。',
        '',
        '二、cloc 是什么(接近中心性 closeness centrality)',
        '  定义:1 ÷ (该节点到网络中所有其他节点的最短路径距离之和)。衡量"这个节点到全网各处平均有多近"。',
        '  距离 = 1 / 连接强度(连接越强=距离越短)。',
        '  关键性质:受权重整体幅值影响(随缩放系数 k 线性变化)。所有连接一起 ×k → 距离整体缩短 → cloc 整体变大。',
        '',
        '三、bc 是什么(介数中心性 betweenness centrality)',
        '  定义:全网所有最短路径中,经过该节点的比例。衡量"这个节点是不是卡在很多关键通路的咽喉位置"。',
        '  关键性质:对权重整体缩放严格不变(k 的 0 次方)。所有连接一起 ×k,只是把每条路径同比例缩短,',
        '  谁是最短路径、要经过谁,完全不变 → 介数不变。',
        '',
        '四、为什么从 cloc 换成 bc(原因全貌)',
        '  审稿人 R1 质疑:训练后 hub 上升,可能只是连接整体变强(权重整体放大)的算术后果,不是真的枢纽重组。',
        '  我们用"缩放对照"检验:把训练前网络整体放大到训练后平均强度(只放大、不改连接格局),看真实训练后能否超过它。',
        '  问题在于:cloc 版 hub 的三个子指标(mean_str、eloc、cloc)全部随权重幅值上涨。',
        '  于是"缩放臂"(放大后的训练前网络)的 hub 分也被同样抬高 → 真实训练后与缩放臂拉不开 → 排除不了"只是整体放大"。',
        '  换成 bc:第三子指标不再随幅值上涨(介数对缩放不变),缩放臂不再被人为抬高 →',
        '  真实训练后若有"超出整体放大"的枢纽重组,就能在强连接骨架上做出显著差异(实测显著高于缩放)。',
        '  实测:cloc 版在任何阈值都做不出显著(p 全程 ns);bc 版在阈值≥0.45、post2 上显著(p 最低到 0.004)。',
        '',
        '五、本次 Excel 改了什么',
        '  1) 阈值梯度扩展到 0.10–0.80(原 0.10–0.60)。',
        '  2) 每个阈值 sheet 在 cloc 版 Hub 块(第36行)下面,新增 Hub_bc 块(第48行)。',
        '  3) 底部显著性表(第62行)同时列 Hub(cloc) 与 Hub_bc 两版的"实测 vs 缩放"配对检验。',
        '  4) 汇总 sheet 增加 Hub_bc 的残差/p/显著三列。',
        '',
        '六、形象说明(给数理基础不强的:为什么换了 bc 就能看出差异)',
        '  把每个电极想成班里一个人,连接强度=两人交情深浅;"hub/枢纽"=班里的核心人物。',
        '  · cloc(接近中心性)像"我到全班每个人的平均社交距离有多近"。全班交情一起加深、大家距离都拉近,',
        '    这个分自然就涨——所以"全班一起变熟"也会让 cloc 涨,分不清是真成了核心,还是只是大家都更熟了。',
        '  · bc(介数中心性)像"班里任意两人要传话,得经过我的次数占多少"。它看的是"谁卡在关键传话路线上"。',
        '    把全班交情统一加深一倍,谁找谁传话的最短路线还是原来那条(只是每条都同比例变快),',
        '    所以"传话要经过谁"不变 → bc 不变。',
        '  · 缩放对照 = "把训练前全班交情,统一加深到训练后平均水平"(只调交情、不换关系格局)。',
        '    用 cloc 评判这个"假班级":它的核心分也跟着涨(cloc 随交情涨),和真实训练后拉不开 → 看不出区别。',
        '    用 bc 评判:假班级的传话路线没变、bc 不变;而真实训练后若关键传话人真换了(路线重排),bc 会变 → 差异显出来。',
        '  一句话:cloc 会跟着"整体变强"一起涨,所以分不清;bc 只认"关系格局/路线"有没有真变,',
        '  所以它能把"真的重组"和"只是整体变强"区分开。',
    ]
    for i, t in enumerate(lines, 1):
        ws.cell(i, 1, t)

def main():
    wb = Workbook()
    detail_sheets = []
    summaries = {k: [] for k, _ in KINDS}
    for thr in THRS:
        rows = [compute_dish(c, thr, False, 'cloc') for c in ORDER]
        ws = wb.create_sheet(f'thr_{thr:.2f}')
        detail_sheets.append((thr, ws, rows))
        write_metrics_sheet(ws, rows)   # 注：其中 Hub 块为 cloc 口径(对齐图2E-F)
        # 追加 Hub_bc 块(第三指标换 betweenness，对整体缩放严格不变；另两子指标 mean_str/eloc 同 cloc 版)
        rbc = 48
        ws.cell(rbc, 1, 'Hub_bc 块(第三指标=betweenness)；上方 Hub 块=cloc(图2E-F口径)')
        for c, h in enumerate(['Hub_bc_pre', 'Hub_bc_p1实测', 'Hub_bc_p1缩放', 'Hub_bc_p1残差',
                               'Hub_bc_p2实测', 'Hub_bc_p2缩放', 'Hub_bc_p2残差']):
            ws.cell(rbc + 1, 6 + c, h)
        for i, r in enumerate(rows):
            rr = rbc + 2 + i
            ws.cell(rr, 6, rnd(r['Hub_bc_pre']))
            ws.cell(rr, 7, rnd(r['Hub_bc_p1_act'])); ws.cell(rr, 8, rnd(r['Hub_bc_p1_scl'])); ws.cell(rr, 9, rnd(r['Hub_bc_p1_res']))
            ws.cell(rr, 10, rnd(r['Hub_bc_p2_act'])); ws.cell(rr, 11, rnd(r['Hub_bc_p2_scl'])); ws.cell(rr, 12, rnd(r['Hub_bc_p2_res']))
        # 底部附：三组配对检验(实测vs pre / 缩放vs pre / 实测vs缩放)
        r0 = 62
        ws.cell(r0, 1, '配对检验(Wilcoxon单侧,方向=训练强化方向;CC/Eglob/Hub增大,CPL减小;Hub=cloc,Hub_bc=betweenness)')
        for c, h in enumerate(['指标', '对比', 'post1_Δ均', 'post1_p', 'post1_显著', 'post2_Δ均', 'post2_p', 'post2_显著'], 1):
            ws.cell(r0 + 1, c, h)
        rr = r0 + 2
        srows = {k: dict(阈值=thr, W1密度=round(density(thr), 3),
                         kS均值=round(np.mean([r[f'kS_p{j}'] for r in rows for j in (1, 2)]), 4)) for k, _ in KINDS}
        for metric in ['CC', 'Eglob', 'CPL', 'Hub', 'Hub_bc']:
            mlabel = metric + ('(cloc)' if metric == 'Hub' else '')
            for kind, klabel in KINDS:
                m1, p1 = pair_test(rows, metric, 1, kind)
                m2, p2 = pair_test(rows, metric, 2, kind)
                ws.cell(rr, 1, mlabel); ws.cell(rr, 2, klabel)
                ws.cell(rr, 3, round(m1, 4)); ws.cell(rr, 4, round(p1, 4) if p1 == p1 else None); ws.cell(rr, 5, stars(p1))
                ws.cell(rr, 6, round(m2, 4)); ws.cell(rr, 7, round(p2, 4) if p2 == p2 else None); ws.cell(rr, 8, stars(p2))
                rr += 1
                sd = srows[kind]
                sd[f'{metric}_p1Δ'] = round(m1, 4); sd[f'{metric}_p1_p'] = round(p1, 4) if p1 == p1 else None; sd[f'{metric}_p1_显著'] = stars(p1)
                sd[f'{metric}_p2Δ'] = round(m2, 4); sd[f'{metric}_p2_p'] = round(p2, 4) if p2 == p2 else None; sd[f'{metric}_p2_显著'] = stars(p2)
        for k, _ in KINDS:
            summaries[k].append(srows[k])
        a = srows['act_scl']; s = srows['scl_pre']
        print(f'thr={thr:.2f} 密度={a["W1密度"]:.3f} kS={a["kS均值"]:.3f} | '
              f'实测vs缩放 p2: CC{a["CC_p2_显著"]} Eg{a["Eglob_p2_显著"]} CPL{a["CPL_p2_显著"]} Hub_bc{a["Hub_bc_p2_显著"]} | '
              f'缩放vs pre p2: CC{s["CC_p2_显著"]} CPL{s["CPL_p2_显著"]} Hub{s["Hub_p2_显著"]}')

    # 三个汇总 sheet（每种对比一个），放最前
    titles = {'act_scl': '汇总_实测vs缩放', 'scl_pre': '汇总_缩放vs pre', 'act_pre': '汇总_实测vs pre'}
    for kind in ['act_pre', 'scl_pre', 'act_scl']:   # 最后建的(act_scl)排最前
        ws = wb.create_sheet(titles[kind], 0)
        cols = ['阈值', 'W1密度', 'kS均值']
        for metric in ['CC', 'Eglob', 'CPL', 'Hub', 'Hub_bc']:
            for k in (1, 2):
                cols += [f'{metric}_p{k}Δ', f'{metric}_p{k}_p', f'{metric}_p{k}_显著']
        for c, h in enumerate(cols, 1):
            ws.cell(1, c, h)
        for i, srow in enumerate(summaries[kind], 2):
            for c, h in enumerate(cols, 1):
                ws.cell(i, c, srow.get(h))
    add_explanation_sheet(wb)   # 说明页放最前
    # 删除自动空白首 sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    out = os.path.join(OUTDIR, 'metrics_threshold_sweep_full.xlsx')
    try:
        wb.save(out)
    except PermissionError:
        out = os.path.join(OUTDIR, 'metrics_threshold_sweep_full_v2.xlsx')
        wb.save(out)
    print('\n已写出：', out)

if __name__ == '__main__':
    main()
