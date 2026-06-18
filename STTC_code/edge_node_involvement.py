# edge_node_involvement.py
# 用途与使用方法：
#   回应 R1[4]（节点层面）。把边层面四类连边变化（增强/减弱/新增/修剪）按电极拆开，
#   逐电极统计它参与的各类变化连边数（= 附图 1 中节点大小的底层数据），并给出多层面的参与占比与集中度，
#   作为"节点层面重组不依赖权重幅值"的定量证据，也供成图（Lorenz/柱状/空间图）取数。
#   运行：python edge_node_involvement.py
# 输入：
#   复用 param_sweep_analysis 的 load_full_symm / DESIGNATION / ORDER / MERGED；数据 = <merged>\<盘>\sttc\sttc-spikes_spon<s>.mat 的 adjM。
#   四类判定口径与 main_edge_change.m / classify_edges 一致：存在阈 τ=0.35、变化阈 δ=0.10、仅判 pre∩post 公共活跃电极对。
#   两个递进阶段：pre→post1、pre→post2。
# 输出（写到 <merged>\param_sweep\edge_node_involvement.xlsx）：
#   1) 逐电极明细  —— 每(盘×阶段×电极)：增强/减弱/新增/修剪数、变化合计、活跃边数、卷入率、卷入占比、排名
#   2) 逐盘集中度  —— 每(盘×阶段)：活跃电极数、卷入电极占比、总卷入量、人均/中位/最大、前10/25/50%电极占比、基尼系数
#   3) 合计均值    —— 9 盘按阶段汇总(均值)
#   4) 说明        —— 口径与各列定义(中文工作版；投稿附件需另出英文表头版)
# 数据流：sttc .mat(adjM) → load_full_symm → classify 四类 → 按电极累加 → 占比/集中度 → Excel
import os
import numpy as np
from openpyxl import Workbook
from param_sweep_analysis import load_full_symm, DESIGNATION, ORDER, OUTDIR

TAU, DELTA = 0.35, 0.10
CATS = ['增强', '减弱', '新增', '修剪']


def classify_per_node(Wpre, Wpost, common):
    """返回每电极各类计数 dict + 活跃边数(并集度) + 稳定数。索引对齐 common。"""
    idx = np.array(common, int) - 1
    a = Wpre[np.ix_(idx, idx)]
    b = Wpost[np.ix_(idx, idx)]
    m = idx.size
    cnt = {c: np.zeros(m, int) for c in CATS}
    stable = np.zeros(m, int)
    deg_union = np.zeros(m, int)   # pre 或 post 中 ≥τ 的连边数（该电极的连接规模）
    for ii in range(m - 1):
        for jj in range(ii + 1, m):
            av, bv = a[ii, jj], b[ii, jj]
            ppre, ppost = av >= TAU, bv >= TAU
            if ppre or ppost:
                deg_union[ii] += 1
                deg_union[jj] += 1
            if not ppre and not ppost:
                continue
            cat = None
            if ppre and ppost:
                if bv - av > DELTA:
                    cat = '增强'
                elif av - bv > DELTA:
                    cat = '减弱'
                else:
                    stable[ii] += 1
                    stable[jj] += 1
            elif (not ppre) and ppost and bv >= TAU + DELTA:
                cat = '新增'
            elif ppre and (not ppost) and av >= TAU + DELTA:
                cat = '修剪'
            else:
                stable[ii] += 1
                stable[jj] += 1
            if cat is not None:
                cnt[cat][ii] += 1
                cnt[cat][jj] += 1
    return cnt, deg_union, stable


def gini(x):
    x = np.sort(np.asarray(x, float))
    n = x.size
    s = x.sum()
    if n == 0 or s == 0:
        return 0.0
    idx = np.arange(1, n + 1)
    return (2.0 * (idx * x).sum()) / (n * s) - (n + 1.0) / n


def top_share(vals, frac):
    v = np.sort(np.asarray(vals, float))[::-1]
    s = v.sum()
    if s == 0:
        return 0.0
    k = max(1, int(round(frac * v.size)))
    return 100.0 * v[:k].sum() / s


def r(x, n=1):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    return round(float(x), n)


def main():
    detail = []     # 逐电极
    conc = []       # 逐盘集中度
    for cell in ORDER:
        pre, posts = DESIGNATION[cell]
        Wpre, actPre = load_full_symm(cell, pre)
        for k, t in enumerate(posts, 1):
            stage = f'post{k}'
            Wpost, actPost = load_full_symm(cell, t)
            common = np.intersect1d(actPre, actPost)
            common = common[(common >= 1) & (common <= Wpre.shape[0])]
            cnt, deg_union, stable = classify_per_node(Wpre, Wpost, common)
            changed = sum(cnt[c] for c in CATS)        # 每电极变化合计(向量)
            total_invol = changed.sum()                # 全盘总卷入量(=2×变化边数)
            m = common.size
            order = np.argsort(changed)[::-1]
            rank = np.empty(m, int)
            rank[order] = np.arange(1, m + 1)
            for i in range(m):
                detail.append(dict(
                    盘=cell, 阶段=stage, 电极号=int(common[i]),
                    增强=int(cnt['增强'][i]), 减弱=int(cnt['减弱'][i]),
                    新增=int(cnt['新增'][i]), 修剪=int(cnt['修剪'][i]),
                    变化合计=int(changed[i]), 稳定=int(stable[i]), 活跃边数=int(deg_union[i]),
                    卷入率=r(changed[i] / deg_union[i] if deg_union[i] > 0 else 0.0, 3),
                    卷入占比百分比=r(100.0 * changed[i] / total_invol if total_invol > 0 else 0.0, 2),
                    卷入排名=int(rank[i]),
                ))
            n_inv = int((changed > 0).sum())
            conc.append(dict(
                盘=cell, 阶段=stage, 活跃电极数=m, 卷入电极数=n_inv,
                卷入电极占比百分比=r(100.0 * n_inv / m, 1),
                变化连边数=int(total_invol // 2), 总卷入量=int(total_invol),
                人均卷入=r(changed.mean(), 2), 卷入中位=int(np.median(changed)), 卷入最大=int(changed.max()),
                前10pct电极占比=r(top_share(changed, 0.10), 1),
                前25pct电极占比=r(top_share(changed, 0.25), 1),
                前50pct电极占比=r(top_share(changed, 0.50), 1),
                基尼系数=r(gini(changed), 3),
            ))

    wb = Workbook()

    # Sheet 1 逐电极明细
    ws = wb.active
    ws.title = '逐电极明细'
    cols1 = ['盘', '阶段', '电极号', '增强', '减弱', '新增', '修剪', '变化合计', '稳定', '活跃边数',
             '卷入率', '卷入占比百分比', '卷入排名']
    ws.append(cols1)
    for d in detail:
        ws.append([d[c] for c in cols1])

    # Sheet 2 逐盘集中度
    ws2 = wb.create_sheet('逐盘集中度')
    cols2 = ['盘', '阶段', '活跃电极数', '卷入电极数', '卷入电极占比百分比', '变化连边数', '总卷入量',
             '人均卷入', '卷入中位', '卷入最大', '前10pct电极占比', '前25pct电极占比', '前50pct电极占比', '基尼系数']
    ws2.append(cols2)
    for d in conc:
        ws2.append([d[c] for c in cols2])

    # Sheet 3 合计均值（按阶段）
    ws3 = wb.create_sheet('合计均值')
    ws3.append(['说明：9 盘按阶段取均值；均匀参与下 前10/25/50%占比应≈10/25/50%、基尼≈0'])
    ws3.append(cols2[:2] + cols2[4:])
    for stage in ('post1', 'post2'):
        rows = [d for d in conc if d['阶段'] == stage]
        avg = {c: np.mean([d[c] for d in rows]) for c in cols2[4:]}
        ws3.append(['9盘均值', stage] + [r(avg[c], 2) for c in cols2[4:]])

    # Sheet 4 说明
    ws4 = wb.create_sheet('说明')
    ws4.column_dimensions['A'].width = 120
    for line in [
        '【电极卷入度 — 全部列定义与计算方法（逐电极明细 / 逐盘集中度 / 合计均值）】',
        '',
        '〔通用判定规则（三张表共用）〕',
        '  数据：两个递进阶段 post1=pre→post1（train1 效果）、post2=pre→post2（累积效果）；只判训练前后都活跃(common-active)的电极对。',
        '  存在阈 τ=0.35：STTC≥0.35 才算连接存在；变化阈 δ=0.10：权重变化>0.10 才算增强/减弱；新增/修剪需跨 τ+δ=0.45（防近阈抖动）。',
        '  一条变化连边记到它的两个端点电极上，故全盘“总卷入量”=2×变化连边数。',
        '  四类判定与边层面图 2G 完全一致；本表只是把“全盘各类总数”拆到每个电极上数。',
        '',
        '〔Sheet 1 · 逐电极明细 — 每(盘×阶段×电极)一行〕',
        '  盘：培养盘编号。',
        '  阶段：post1 或 post2（见上）。',
        '  电极号：MEA 电极/节点编号（仅列 common-active 电极）。',
        '  增强：该电极参与的变强连接数。某伙伴 j 的 pre、post 都≥0.35 且 post−pre>0.10 → 计1。',
        '  减弱：pre、post 都≥0.35 且 pre−post>0.10 → 计1。',
        '  新增：pre<0.35（原不存在）、post≥0.45 → 计1。',
        '  修剪：pre≥0.45、post<0.35 → 计1。',
        '  变化合计：增强+减弱+新增+修剪（=附图 1 中该电极的节点大小）。',
        '  稳定：两端都≥0.35 且 |post−pre|≤0.10 的连接数（存在但没明显变化）。',
        '  活跃边数：该电极的连接规模 = pre 或 post 中与它≥0.35 的伙伴电极数（并集）。',
        '  卷入率：变化合计 ÷ 活跃边数 = 它自己的连接里有多大比例被重连（已按连接规模归一）。',
        '  卷入占比百分比：100×变化合计 ÷ 全盘总卷入量 = 它占全盘所有重连的份额。',
        '  卷入排名：本盘内按变化合计降序的名次，1=参与最多。',
        '',
        '〔Sheet 2 · 逐盘集中度 — 每(盘×阶段)一行，判断卷入是否集中在少数电极〕',
        '  盘 / 阶段：同上。',
        '  活跃电极数：该盘 common-active 电极总数。',
        '  卷入电极数：变化合计>0（至少参与1条变化连边）的电极数。',
        '  卷入电极占比百分比：100×卷入电极数 ÷ 活跃电极数。',
        '  变化连边数：该盘四类变化连边总条数。',
        '  总卷入量：所有电极“变化合计”之和 = 2×变化连边数。',
        '  人均卷入：总卷入量 ÷ 活跃电极数（每个电极平均碰几条变化边）。',
        '  卷入中位：各电极“变化合计”的中位数。',
        '  卷入最大：卷入最多的电极的“变化合计”。',
        '  前10pct/前25pct/前50pct电极占比：卷入度最高的前 N% 电极承担的卷入量占全盘比例。算法：电极按变化合计降序，取前 round(N%×电极数) 个，其卷入量之和 ÷ 总卷入量 ×100。均匀时≈10/25/50，越高越集中。',
        '  基尼系数：卷入度分布的不均匀度，0=完全均匀、越大越集中。算法 G=(2·Σ i·x(i))/(n·Σx) − (n+1)/n，x(i) 为升序排序后的各电极卷入度，i=1..n。',
        '',
        '〔Sheet 3 · 合计均值 — 9 盘按阶段(post1/post2)取均值〕',
        '  列名与 Sheet 2 相同（卷入电极占比百分比、变化连边数、总卷入量、人均卷入、卷入中位、卷入最大、前10/25/50pct电极占比、基尼系数）；数值为 9 个盘在该阶段的算术平均。',
        '  用途：回复正文引用的“头条数”，如 post2 前25%占比≈56%、基尼≈0.44。',
        '  参照基准：均匀参与下 前10/25/50%占比应≈10/25/50、基尼≈0。',
        '',
        '〔用途与边界〕',
        '  该计数基于连接的类别变化、不随权重整体幅值改变，是节点层面重组的幅值无关刻画（回应 R1[4]）。',
        '  本表中文表头为内部工作版；作为投稿附件需另出英文表头版。',
    ]:
        ws4.append([line])

    out = os.path.join(OUTDIR, 'edge_node_involvement.xlsx')
    try:
        wb.save(out)
    except PermissionError:
        out = os.path.join(OUTDIR, 'edge_node_involvement_v2.xlsx')
        wb.save(out)

    # 控制台核验
    print('=== 逐盘集中度（前25%电极占比 / 基尼，post2）===')
    for d in conc:
        if d['阶段'] == 'post2':
            print(f"  {d['盘']}: 卷入电极{d['卷入电极占比百分比']}%  前25%占{d['前25pct电极占比']}%  基尼{d['基尼系数']}  最大卷入{d['卷入最大']}")
    p2 = [d for d in conc if d['阶段'] == 'post2']
    print(f"\n9盘均值 post2: 前10%占{np.mean([d['前10pct电极占比'] for d in p2]):.0f}%  "
          f"前25%占{np.mean([d['前25pct电极占比'] for d in p2]):.0f}%  "
          f"前50%占{np.mean([d['前50pct电极占比'] for d in p2]):.0f}%  基尼{np.mean([d['基尼系数'] for d in p2]):.2f}")
    print('已写出：', out)


if __name__ == '__main__':
    main()
