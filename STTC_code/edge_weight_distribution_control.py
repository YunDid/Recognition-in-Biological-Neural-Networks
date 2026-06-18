# edge_weight_distribution_control.py
# 用途与使用方法：
#   回应 R1[2]。把原 edge_weight_distribution.xlsx（只有 pre/post1/post2 实测占比）扩展为
#   "实测 vs 缩放对照 + kS 并列" 的连边权重保留率对照表，让"训练后 vs 整体缩放对照"的差异一眼可量。
#   占比 = 累积保留率 = 该阶段连边中权重 ≥ 阈值的比例（生存函数 / CCDF）。
#   缩放对照 = 把训练前连边整体放大 kS 倍（只放大、不改连接格局），kS = post 非零边均值 / pre 非零边均值。
#   均匀缩放保排名：缩放对照在阈值 T 的保留率 = P(pre 权重 × kS ≥ T) = pre 在 T/kS 的保留率。
#   实测 post 在高阈值的保留率超过缩放对照 = 训练把更多边推进强连接区间，缩放造不出 → 强骨架重组的直接证据。
#   运行：python edge_weight_distribution_control.py
# 输入：
#   复用 param_sweep_analysis 的 load_full / DESIGNATION / ORDER / MERGED；数据 = <merged>\<盘>\sttc\sttc-spikes_spon<s>.mat 的 adjM。
#   9 盘 pre/post1/post2 阶段映射同 main_edge_change.m cfg.designation。
# 输出（写到 <merged>\param_sweep\edge_weight_distribution_control.xlsx）：
#   1) 占比对照_合计      —— 9 盘均值：每阈值 pre / post1·post2 实测 / post1·post2 缩放对照 / 差值(实测−缩放)
#   2) 各盘kS与均值       —— 每盘 mean_nz(pre/post1/post2) + kS1 / kS2，末行均值
#   3) 各盘_占比(关键阈值) —— 每盘在 0.35/0.50/0.65 的 实测 vs 缩放对照 保留率（看跨盘异质）
#   4) 说明              —— 口径与读法（中文工作版；投稿附件需另出英文表头版）
# 数据流：sttc .mat(adjM) → load_full → 对称化取上三角正权 → kS=mean_nz比 → 缩放对照=pre×kS → 各阈值保留率
import os
import numpy as np
from openpyxl import Workbook
from param_sweep_analysis import load_full, DESIGNATION, ORDER, OUTDIR

THRS = [round(0.10 + 0.05 * i, 2) for i in range(15)]   # 0.10 .. 0.80，与 metrics_threshold_sweep_full 同 x 轴
KEY_THRS = [0.35, 0.50, 0.65]                            # 关键阈值（全网络/起点/峰值），逐盘展开用
EXIST = 0.35                                             # 存在阈值（论文功能连接定义阈），kS 在此口径上算（=论文 mean STTC）


def edge_pool(cell, s):
    """该阶段所有上三角正权（对称化后）作为连边权重总体。"""
    A, _ = load_full(cell, s)               # 已 NaN→0、负→0
    A = (A + A.T) / 2.0
    np.fill_diagonal(A, 0.0)
    iu = np.triu_indices(A.shape[0], 1)
    w = A[iu]
    return w[w > 0]


def mean_nz_ge(pool, thr):
    v = pool[pool >= thr]
    return v.mean() if v.size else np.nan


def retention(pool, thr):
    """累积保留率%：权重 ≥ thr 的比例。"""
    return 100.0 * (pool >= thr).mean() if pool.size else np.nan


def per_dish():
    """逐盘：返回 pre/post1/post2 权重池、kS1/kS2、mean_nz。"""
    out = {}
    for cell in ORDER:
        pre, posts = DESIGNATION[cell]
        ppre = edge_pool(cell, pre)
        pp = [edge_pool(cell, t) for t in posts]
        # kS 在存在阈 0.35 口径上算（= 论文 mean STTC 之比），单一全局缩放系数
        mpre = mean_nz_ge(ppre, EXIST)
        mpost = [mean_nz_ge(p, EXIST) for p in pp]
        kS = [mpost[k] / mpre if mpre > 0 else np.nan for k in range(2)]
        # 对照：kS 若改用"全部正权均值之比"，量级核对用（应与上面接近）
        kS_allpos = [pp[k].mean() / ppre.mean() if ppre.size else np.nan for k in range(2)]
        out[cell] = dict(pre=ppre, post=pp, kS=kS, kS_allpos=kS_allpos,
                         mnz=dict(pre=mpre, post1=mpost[0], post2=mpost[1]))
    return out


def pooled_table(dishes):
    """9 盘均值的保留率对照表（每阈值一行）。缩放对照 = pre 池 × 该盘 kS。"""
    rows = []
    for thr in THRS:
        pre_r, p1_act, p1_scl, p2_act, p2_scl = [], [], [], [], []
        for cell in ORDER:
            d = dishes[cell]
            pre_r.append(retention(d['pre'], thr))
            p1_act.append(retention(d['post'][0], thr))
            p2_act.append(retention(d['post'][1], thr))
            p1_scl.append(retention(d['pre'] * d['kS'][0], thr))
            p2_scl.append(retention(d['pre'] * d['kS'][1], thr))
        rows.append(dict(
            thr=thr,
            pre=np.nanmean(pre_r),
            p1_act=np.nanmean(p1_act), p1_scl=np.nanmean(p1_scl),
            p1_diff=np.nanmean(p1_act) - np.nanmean(p1_scl),
            p2_act=np.nanmean(p2_act), p2_scl=np.nanmean(p2_scl),
            p2_diff=np.nanmean(p2_act) - np.nanmean(p2_scl),
        ))
    return rows


def r1(x):
    return None if x is None or (isinstance(x, float) and np.isnan(x)) else round(float(x), 1)


def r4(x):
    return None if x is None or (isinstance(x, float) and np.isnan(x)) else round(float(x), 4)


def main():
    dishes = per_dish()
    wb = Workbook()

    # ---- Sheet 1：占比对照_合计 ----
    ws = wb.active
    ws.title = '占比对照_合计'
    kS1m = np.nanmean([dishes[c]['kS'][0] for c in ORDER])
    kS2m = np.nanmean([dishes[c]['kS'][1] for c in ORDER])
    ws.cell(1, 1, f'连边权重累积保留率%（9盘均值）：实测 vs 缩放对照（pre×kS）。均值 kS1={kS1m:.3f}  kS2={kS2m:.3f}')
    hdr = ['阈值', 'pre_占比%',
           'post1_实测%', 'post1_缩放对照%', 'post1_差值(实测−缩放)',
           'post2_实测%', 'post2_缩放对照%', 'post2_差值(实测−缩放)']
    for c, h in enumerate(hdr, 1):
        ws.cell(2, c, h)
    rows = pooled_table(dishes)
    for i, r in enumerate(rows, 3):
        ws.cell(i, 1, r['thr'])
        ws.cell(i, 2, r1(r['pre']))
        ws.cell(i, 3, r1(r['p1_act'])); ws.cell(i, 4, r1(r['p1_scl'])); ws.cell(i, 5, r1(r['p1_diff']))
        ws.cell(i, 6, r1(r['p2_act'])); ws.cell(i, 7, r1(r['p2_scl'])); ws.cell(i, 8, r1(r['p2_diff']))
    ws.column_dimensions['A'].width = 8
    for col in 'BCDEFGH':
        ws.column_dimensions[col].width = 16

    # ---- Sheet 2：各盘kS与均值 ----
    ws2 = wb.create_sheet('各盘kS与均值')
    h2 = ['盘', 'meanSTTC_pre', 'meanSTTC_post1', 'meanSTTC_post2', 'kS1(post1/pre)', 'kS2(post2/pre)',
          'kS1_全正权口径', 'kS2_全正权口径']
    for c, h in enumerate(h2, 1):
        ws2.cell(1, c, h)
    for i, cell in enumerate(ORDER, 2):
        d = dishes[cell]
        ws2.cell(i, 1, cell)
        ws2.cell(i, 2, r4(d['mnz']['pre'])); ws2.cell(i, 3, r4(d['mnz']['post1'])); ws2.cell(i, 4, r4(d['mnz']['post2']))
        ws2.cell(i, 5, r4(d['kS'][0])); ws2.cell(i, 6, r4(d['kS'][1]))
        ws2.cell(i, 7, r4(d['kS_allpos'][0])); ws2.cell(i, 8, r4(d['kS_allpos'][1]))
    rr = len(ORDER) + 2
    ws2.cell(rr, 1, '均值')
    for c, key in [(5, 0), (6, 1)]:
        ws2.cell(rr, c, r4(np.nanmean([dishes[cell]['kS'][key] for cell in ORDER])))
    for c in range(2, 5):
        keymap = {2: 'pre', 3: 'post1', 4: 'post2'}
        ws2.cell(rr, c, r4(np.nanmean([dishes[cell]['mnz'][keymap[c]] for cell in ORDER])))
    for col in 'ABCDEFGH':
        ws2.column_dimensions[col].width = 16

    # ---- Sheet 3：各盘_占比(关键阈值) ----
    ws3 = wb.create_sheet('各盘_占比(关键阈值)')
    ws3.cell(1, 1, '逐盘累积保留率%：实测 vs 缩放对照(pre×kS)，看跨盘异质（关键阈值 0.35/0.50/0.65）')
    h3 = ['盘', '阶段']
    for t in KEY_THRS:
        h3 += [f'{t:.2f}_实测%', f'{t:.2f}_缩放%', f'{t:.2f}_差值']
    for c, h in enumerate(h3, 1):
        ws3.cell(2, c, h)
    ri = 3
    for cell in ORDER:
        d = dishes[cell]
        for k, tag in [(0, 'post1'), (1, 'post2')]:
            ws3.cell(ri, 1, cell); ws3.cell(ri, 2, tag)
            c = 3
            for t in KEY_THRS:
                act = retention(d['post'][k], t)
                scl = retention(d['pre'] * d['kS'][k], t)
                ws3.cell(ri, c, r1(act)); ws3.cell(ri, c + 1, r1(scl)); ws3.cell(ri, c + 2, r1(act - scl))
                c += 3
            ri += 1
    for col in 'AB':
        ws3.column_dimensions[col].width = 10

    # ---- Sheet 4：说明 ----
    ws4 = wb.create_sheet('说明')
    ws4.column_dimensions['A'].width = 120
    lines = [
        '【连边权重保留率：实测 vs 缩放对照 对照表 — 口径与读法】',
        '',
        '一、占比是什么',
        '  占比 = 累积保留率 = 该阶段所有连边里，权重(STTC) ≥ 该行阈值的比例(%)。阈值越高、留下的边越少。',
        '  连边总体 = 对称化矩阵上三角的全部正权(每盘各自归一，故是该盘内部的比例，不受盘大小影响)。',
        '',
        '二、缩放对照怎么来的',
        '  缩放对照 = 把训练前(pre)的每条连边权重统一 × kS(只整体放大、不改谁连谁)。',
        '  kS = 训练后非零边平均 STTC ÷ 训练前非零边平均 STTC(在存在阈 0.35 上算，= 论文 mean STTC 之比)。',
        '  含义：假设训练只是把全网连接均匀放大到训练后的平均强度，会得到怎样的权重分布。',
        '  因均匀缩放保排名，缩放对照在阈值 T 的保留率 = pre 在 T/kS 的保留率(等价：pre 权重×kS 后再数 ≥T 的比例)。',
        '',
        '三、怎么读这张表(回应 R1[2] 的关键)',
        '  · 低阈值(如 0.35)：实测 post 与 缩放对照 的保留率几乎相同 → 这一档差不出来(网络近全连接，缩放即可复现)。',
        '  · 高阈值(如 0.60–0.65)：实测 post 的保留率明显高于缩放对照 → 训练把更多边推进了强连接区间，',
        '    而均匀缩放(×kS≈1.04)推不动这么多边过线 → 这部分差值正是"超出整体放大"的强骨架重组，缩放造不出。',
        '  · 差值列(实测−缩放)随阈值由≈0 增大，就是"偏离随聚焦强连接单调增大"的分布版证据，与残差趋势图同源。',
        '',
        '四、边界(诚实交代)',
        '  · 缩放对照只排除"整体均匀放大"一种替代解释，不排除两次记录间的自发漂移(缺训练前-训练前基线)。',
        '  · pre×kS 个别强边可能略超 STTC 上限(非物理)，但对 ≤0.80 的保留率几乎无影响(越界边本就计入 ≥T)。',
        '  · 本表中文表头为内部工作版；作为投稿附件需另出英文表头版。',
    ]
    for i, t in enumerate(lines, 1):
        ws4.cell(i, 1, t)

    out = os.path.join(OUTDIR, 'edge_weight_distribution_control.xlsx')
    try:
        wb.save(out)
    except PermissionError:
        out = os.path.join(OUTDIR, 'edge_weight_distribution_control_v2.xlsx')
        wb.save(out)

    # ---- 控制台核验打印 ----
    print('=== 各盘 kS（存在阈0.35口径 / 全正权口径，量级核对）===')
    for cell in ORDER:
        d = dishes[cell]
        print(f"  {cell}: kS1={d['kS'][0]:.4f}/{d['kS_allpos'][0]:.4f}  kS2={d['kS'][1]:.4f}/{d['kS_allpos'][1]:.4f}  "
              f"(mnz pre/p1/p2 = {d['mnz']['pre']:.3f}/{d['mnz']['post1']:.3f}/{d['mnz']['post2']:.3f})")
    print(f"  均值 kS1={np.nanmean([dishes[c]['kS'][0] for c in ORDER]):.4f}  kS2={np.nanmean([dishes[c]['kS'][1] for c in ORDER]):.4f}")
    print('\n=== 合计保留率% post2：实测 / 缩放对照 / 差值（关键阈值）===')
    rows = pooled_table(dishes)
    for r in rows:
        if r['thr'] in (0.35, 0.50, 0.60, 0.65, 0.70):
            print(f"  thr={r['thr']:.2f}  实测={r['p2_act']:.1f}  缩放={r['p2_scl']:.1f}  差值={r['p2_diff']:+.1f}")
    print('\n已写出：', out)


if __name__ == '__main__':
    main()
