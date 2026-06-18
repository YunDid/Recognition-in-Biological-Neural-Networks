# -*- coding: utf-8 -*-
import openpyxl, sys
from statistics import mean, pstdev, stdev
from scipy.stats import mannwhitneyu
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r'E:\Recognition-in-Biological-Neural-Networks\Data\Spon\burst_metrics.xlsx', data_only=True)

METRICS = ['MBR 平均爆发率(bursts/min)', 'MNBR 网络同步爆发率(NB/min)',
           'SIB 爆发内spike占比(%)', 'NBD 网络爆发时长(s)']

def load(sheet):
    ws = wb[sheet]; H = [c.value for c in ws[1]]
    return H, [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]

def col(H, name): return H.index(name)

def collect(sheet, div, level):
    H, rows = load(sheet)
    di, pi = col(H,'DIV(天)'), col(H,'盘')
    out = {m: [] for m in METRICS}
    if level == 'rec':
        for r in rows:
            if r[di] != div: continue
            for m in METRICS:
                v = r[col(H,m)]
                if isinstance(v,(int,float)) and v==v: out[m].append(v)
    else:  # dish: average reps within each 盘
        dishes = {}
        for r in rows:
            if r[di] != div: continue
            dishes.setdefault(r[pi], []).append(r)
        for dish, rs in dishes.items():
            for m in METRICS:
                vals = [rr[col(H,m)] for rr in rs if isinstance(rr[col(H,m)],(int,float)) and rr[col(H,m)]==rr[col(H,m)]]
                if vals: out[m].append(mean(vals))
    return out

def fmt(x): return f'{mean(x):.2f}±{(stdev(x) if len(x)>1 else 0):.2f}(n{len(x)})' if x else '—'

for level,label in [('rec','【录制级 (每段5min算一条)】'),('dish','【盘级 (每盘先平均reps)】')]:
    print(f'\n===== DIV20 对比 {label} =====')
    print(f'{"指标":<10}{"单脑区":>20}{"多脑区":>20}{"p值(MannWhitney)":>22}')
    S = collect('单脑区',20,level); M = collect('多脑区',20,level)
    for m in METRICS:
        xs, xm = S[m], M[m]
        try:
            p = mannwhitneyu(xs, xm, alternative='two-sided').pvalue if len(xs)>=2 and len(xm)>=2 else float('nan')
        except Exception:
            p = float('nan')
        short = m.split(' ')[0]
        print(f'{short:<10}{fmt(xs):>20}{fmt(xm):>20}{("%.3g"%p) if p==p else "n<2":>22}')
