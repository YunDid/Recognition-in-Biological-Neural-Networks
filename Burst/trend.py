# -*- coding: utf-8 -*-
import openpyxl, sys
from statistics import mean
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r'E:\Recognition-in-Biological-Neural-Networks\Data\Spon\burst_metrics.xlsx', data_only=True)

def load(sheet):
    ws = wb[sheet]; H = [c.value for c in ws[1]]
    keys = ['DIV(天)', 'MBR 平均爆发率(bursts/min)', 'MNBR 网络同步爆发率(NB/min)',
            'SIB 爆发内spike占比(%)', 'NBD 网络爆发时长(s)', '活跃电极数']
    idx = {k: H.index(k) for k in keys}
    return [{k: r[v] for k, v in idx.items()} for r in ws.iter_rows(min_row=2, values_only=True)]

def agg(rows, div, key, silent_zero=False):
    vals = []
    for r in rows:
        if r['DIV(天)'] != div:
            continue
        v = r[key]; act = r['活跃电极数']
        if v is None or (isinstance(v, float) and v != v):
            if silent_zero and act == 0:
                vals.append(0.0)
        else:
            vals.append(v)
    return (f'{round(mean(vals),2)}(n{len(vals)})') if vals else '—'

specs = [('活跃电极数', '活跃电极(全计)', False),
         ('MBR 平均爆发率(bursts/min)', 'MBR(静默记0)', True),
         ('MNBR 网络同步爆发率(NB/min)', 'MNBR(静默记0)', True),
         ('SIB 爆发内spike占比(%)', 'SIB(仅有效)', False),
         ('NBD 网络爆发时长(s)', 'NBD(仅有效)', False)]

for sheet in ['单脑区', '多脑区']:
    rows = load(sheet)
    print(f'\n===== {sheet} =====')
    print(f'{"指标":<14}{"DIV5":>16}{"DIV10":>16}{"DIV15":>16}{"DIV20":>16}')
    for key, short, sz in specs:
        cells = [agg(rows, dv, key, sz) for dv in [5, 10, 15, 20]]
        print(f'{short:<14}' + ''.join(f'{c:>16}' for c in cells))
